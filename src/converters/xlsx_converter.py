"""XLSX (Excel spreadsheet) to Markdown converter."""

import base64
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from pdf2markdown.converters.document_converter import DocumentConverter
from pdf2markdown.core.config import Config
from pdf2markdown.core.models import (
    ConversionMetadata,
    ConversionResult,
    ExtractedImage,
    ExtractedTable,
)

logger = logging.getLogger(__name__)


class XLSXConverter(DocumentConverter):
    """
    Convert XLSX (Excel) files to Markdown.

    Uses pandas for table conversion and openpyxl for charts, images, and metadata.
    Supports multiple sheets, chart extraction, and formula display.

    Dependencies:
        - pandas: Required for reading Excel files and table conversion
        - openpyxl: Required for charts, images, and advanced features
        - Pillow: Required for image processing
    """

    def __init__(self, config: Optional[Config] = None):
        """
        Initialize the XLSX converter.

        Args:
            config: Configuration for conversion options
        """
        super().__init__(config)
        self._pandas_available = PANDAS_AVAILABLE
        self._openpyxl_available = OPENPYXL_AVAILABLE
        self._pil_available = PIL_AVAILABLE

    def convert(self, file_path: Path) -> ConversionResult:
        """
        Convert an XLSX file to Markdown.

        Args:
            file_path: Path to the XLSX file

        Returns:
            ConversionResult containing markdown, images, tables, and metadata

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file is not a valid XLSX file
            Exception: For other conversion errors
        """
        start_time = datetime.now()

        # Validate file
        self.validate_file(file_path)

        logger.info(f"Converting XLSX file: {file_path}")

        # Extract metadata
        metadata = self._extract_metadata(file_path)

        # Read Excel file with pandas
        try:
            if self.config.xlsx_mode == "selected" and self.config.xlsx_sheets:
                # Load only selected sheets
                excel_data = pd.read_excel(file_path, sheet_name=self.config.xlsx_sheets, engine='openpyxl')
            else:
                # Load all sheets
                excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        except Exception as e:
            raise ValueError(f"Failed to read XLSX file: {e}")

        # Extract images and charts if openpyxl is available
        images = []
        if self._openpyxl_available:
            images = self._extract_images_and_charts(file_path)

        # Convert sheets to markdown
        markdown_parts = []
        tables = []

        if isinstance(excel_data, dict):
            # Multiple sheets
            sheet_count = len(excel_data)

            if self.config.xlsx_mode == "combined":
                # Combine all sheets into one document
                markdown_parts.append(f"# Excel Workbook: {file_path.name}\n")
                markdown_parts.append(f"*Contains {sheet_count} sheet(s)*\n\n")

                for sheet_name, df in excel_data.items():
                    markdown_parts.append(f"## Sheet: {sheet_name}\n")

                    sheet_md, sheet_tables = self._convert_dataframe_to_markdown(
                        df, sheet_name, file_path
                    )
                    markdown_parts.append(sheet_md)
                    tables.extend(sheet_tables)
                    markdown_parts.append("\n---\n\n")

            elif self.config.xlsx_mode == "separate":
                # Each sheet as a separate section
                for idx, (sheet_name, df) in enumerate(excel_data.items()):
                    if idx > 0:
                        markdown_parts.append("\n\n")

                    markdown_parts.append(f"# {sheet_name}\n\n")
                    sheet_md, sheet_tables = self._convert_dataframe_to_markdown(
                        df, sheet_name, file_path
                    )
                    markdown_parts.append(sheet_md)
                    tables.extend(sheet_tables)

            else:  # selected
                # Only selected sheets
                for sheet_name, df in excel_data.items():
                    markdown_parts.append(f"# {sheet_name}\n\n")
                    sheet_md, sheet_tables = self._convert_dataframe_to_markdown(
                        df, sheet_name, file_path
                    )
                    markdown_parts.append(sheet_md)
                    tables.extend(sheet_tables)
                    markdown_parts.append("\n\n")
        else:
            # Single sheet (when sheet_name is specified as string)
            markdown_parts.append(f"# {file_path.stem}\n\n")
            sheet_md, sheet_tables = self._convert_dataframe_to_markdown(
                excel_data, "Sheet1", file_path
            )
            markdown_parts.append(sheet_md)
            tables.extend(sheet_tables)

        # Combine markdown
        markdown = "".join(markdown_parts)

        # Add images section if any
        if images and self.config.extract_images:
            markdown += "\n\n## Images and Charts\n\n"
            for img in images:
                if self.config.image_mode == ImageMode.EMBED and img.base64_data:
                    markdown += f"![{img.alt_text}](data:image/{img.format};base64,{img.base64_data})\n\n"
                elif self.config.image_mode == ImageMode.LINK:
                    markdown += f"![{img.alt_text}]({img.path})\n\n"

        # Update metadata
        end_time = datetime.now()
        metadata.conversion_time_seconds = (end_time - start_time).total_seconds()
        metadata.total_images = len(images)
        metadata.total_tables = len(tables)
        metadata.total_words = len(markdown.split())
        metadata.converter_name = self.get_name()

        logger.info(f"XLSX conversion completed in {metadata.conversion_time_seconds:.2f}s")

        return ConversionResult(
            markdown=markdown,
            images=images,
            tables=tables,
            metadata=metadata,
        )

    def _convert_dataframe_to_markdown(
        self,
        df: "pd.DataFrame",
        sheet_name: str,
        source_file: Path
    ) -> Tuple[str, List[ExtractedTable]]:
        """
        Convert a pandas DataFrame to Markdown table.

        Args:
            df: DataFrame to convert
            sheet_name: Name of the sheet
            source_file: Source file path

        Returns:
            Tuple of (markdown string, list of ExtractedTable objects)
        """
        tables = []
        markdown_parts = []

        # Check if DataFrame is empty
        if df.empty:
            return "*Empty sheet*\n\n", []

        # Clean up the DataFrame
        # Replace NaN with empty strings
        df = df.fillna('')

        # Convert to string to handle mixed types
        df = df.astype(str)

        # Check table width
        if len(df.columns) > self.config.xlsx_max_sheet_width:
            # Too wide - use CSV format instead
            markdown_parts.append(f"*Table too wide ({len(df.columns)} columns), showing first {self.config.xlsx_max_sheet_width} columns*\n\n")
            df = df.iloc[:, :self.config.xlsx_max_sheet_width]

        # Convert to markdown table based on format
        try:
            if self.config.table_format == TableFormat.GITHUB:
                table_md = df.to_markdown(index=False)
            elif self.config.table_format == TableFormat.PIPE:
                table_md = df.to_markdown(index=False, tablefmt='pipe')
            elif self.config.table_format == TableFormat.GRID:
                table_md = df.to_markdown(index=False, tablefmt='grid')
            elif self.config.table_format == TableFormat.HTML:
                table_md = df.to_html(index=False)
            else:
                # Default to GitHub flavor
                table_md = df.to_markdown(index=False)

            markdown_parts.append(table_md)
            markdown_parts.append("\n\n")

            # Create ExtractedTable object
            table = ExtractedTable(
                markdown=table_md,
                page=1,  # XLSX doesn't have pages
                bbox=[0, 0, 0, 0],  # No bbox for XLSX
                source=f"{source_file.name}:{sheet_name}",
                rows=len(df),
                columns=len(df.columns),
            )
            tables.append(table)

        except Exception as e:
            logger.warning(f"Failed to convert DataFrame to markdown: {e}")
            markdown_parts.append(f"*Error converting table: {e}*\n\n")

        return "".join(markdown_parts), tables

    def _extract_metadata(self, file_path: Path) -> ConversionMetadata:
        """
        Extract metadata from XLSX file.

        Args:
            file_path: Path to the XLSX file

        Returns:
            ConversionMetadata object
        """
        metadata = ConversionMetadata(
            title=file_path.stem,
            source_file=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            converter_name=self.get_name(),
        )

        # Extract metadata using openpyxl if available
        if self._openpyxl_available:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)

                # Get workbook properties
                props = wb.properties
                if props:
                    metadata.title = props.title or file_path.stem
                    metadata.author = props.creator or None
                    metadata.creator = props.creator or None
                    metadata.subject = props.subject or None
                    metadata.keywords = props.keywords or None

                    if props.created:
                        metadata.creation_date = props.created
                    if props.modified:
                        metadata.modification_date = props.modified

                # Count sheets
                metadata.page_count = len(wb.sheetnames)
                metadata.warnings.append(f"Workbook contains {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}")

                wb.close()

            except Exception as e:
                logger.warning(f"Failed to extract metadata with openpyxl: {e}")
                metadata.warnings.append(f"Could not extract detailed metadata: {e}")

        return metadata

    def _extract_images_and_charts(self, file_path: Path) -> List[ExtractedImage]:
        """
        Extract images and charts from XLSX file using openpyxl.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List of ExtractedImage objects
        """
        images = []

        if not self._openpyxl_available:
            return images

        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)

            image_counter = 1
            chart_counter = 1

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Extract embedded images
                if hasattr(sheet, '_images'):
                    for img in sheet._images:
                        try:
                            image_data = img.ref.getvalue()

                            # Get image properties
                            if self._pil_available:
                                pil_img = PILImage.open(io.BytesIO(image_data))
                                width, height = pil_img.size
                                img_format = pil_img.format.lower() if pil_img.format else 'png'
                            else:
                                width, height = 0, 0
                                img_format = 'png'

                            # Encode as base64 if needed
                            base64_data = None
                            if self.config.image_mode == ImageMode.EMBED:
                                base64_data = base64.b64encode(image_data).decode('utf-8')

                            extracted_img = ExtractedImage(
                                format=img_format,
                                width=width,
                                height=height,
                                page=1,  # XLSX doesn't have pages
                                index=image_counter,
                                path=None,
                                base64_data=base64_data,
                                alt_text=f"Image {image_counter} from sheet '{sheet_name}'",
                                size_bytes=len(image_data),
                            )
                            images.append(extracted_img)
                            image_counter += 1

                        except Exception as e:
                            logger.warning(f"Failed to extract image: {e}")

                # Extract charts if enabled
                if self.config.xlsx_extract_charts and hasattr(sheet, '_charts'):
                    for chart in sheet._charts:
                        try:
                            # Note: openpyxl doesn't provide direct chart image export
                            # We can only provide metadata about the chart
                            chart_info = ExtractedImage(
                                format='chart',
                                width=0,
                                height=0,
                                page=1,
                                index=chart_counter,
                                path=None,
                                base64_data=None,
                                alt_text=f"Chart {chart_counter} from sheet '{sheet_name}': {chart.title.text if chart.title else 'Untitled'}",
                                size_bytes=0,
                            )
                            images.append(chart_info)
                            chart_counter += 1

                        except Exception as e:
                            logger.warning(f"Failed to extract chart info: {e}")

            wb.close()

        except Exception as e:
            logger.warning(f"Failed to extract images/charts: {e}")

        return images

    def supports_ocr(self) -> bool:
        """
        XLSX files don't need OCR support.

        Returns:
            False - XLSX is structured data, not scanned documents
        """
        return False

    def get_name(self) -> str:
        """
        Get the converter name.

        Returns:
            Human-readable converter name
        """
        return "XLSX Converter (pandas + openpyxl)"

    def is_available(self) -> bool:
        """
        Check if required dependencies are installed.

        Returns:
            True if pandas and openpyxl are available
        """
        if not self._pandas_available:
            logger.debug("pandas not available for XLSX conversion")
            return False

        if not self._openpyxl_available:
            logger.debug("openpyxl not available for XLSX conversion")
            return False

        return True

    def get_supported_extensions(self) -> List[str]:
        """
        Get supported file extensions.

        Returns:
            List of supported extensions
        """
        return ['.xlsx', '.xls', '.xlsm']

    def estimate_conversion_time(self, file_path: Path) -> float:
        """
        Estimate conversion time for XLSX file.

        Args:
            file_path: Path to the file

        Returns:
            Estimated time in seconds
        """
        # XLSX conversion is generally fast
        # Estimate based on file size: ~0.5 seconds per MB
        try:
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            return max(1.0, file_size_mb * 0.5)
        except Exception:
            return 2.0  # Default fallback


# Import ImageMode from config for use in this module
from pdf2markdown.core.config import ImageMode, TableFormat
