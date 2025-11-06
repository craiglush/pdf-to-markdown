"""MarkItDown converter - Microsoft's multi-format document to Markdown converter."""

import os
import time
from dataclasses import field
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from pdf2markdown.converters.document_converter import DocumentConverter
from pdf2markdown.core.config import Config
from pdf2markdown.core.models import ConversionMetadata, ConversionResult


class MarkItDownConverter(DocumentConverter):
    """
    Converter using Microsoft MarkItDown for multi-format document conversion.

    Supports 13+ formats:
    - PDF, DOCX, XLSX, PPTX, HTML
    - Images (JPG, PNG, etc.) with OCR and EXIF
    - Audio (WAV, MP3) with transcription
    - YouTube URLs (transcript extraction)
    - EPub, XML, JSON, CSV, ZIP
    - Outlook messages (MSG)

    Features:
    - LLM-powered image descriptions (optional)
    - Azure Document Intelligence for high-accuracy PDFs (optional)
    - Rich metadata extraction (optional)
    - Stream-based processing (no temp files)
    """

    # All supported extensions
    SUPPORTED_EXTENSIONS = [
        # Documents
        '.pdf', '.docx', '.doc', '.xlsx', '.xls', '.pptx', '.ppt',
        # Web
        '.html', '.htm',
        # Images
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp',
        # Audio
        '.wav', '.mp3', '.m4a', '.flac', '.ogg',
        # E-books
        '.epub',
        # Data formats
        '.json', '.xml', '.csv',
        # Archives
        '.zip',
        # Email
        '.msg',
    ]

    def __init__(self, config: Optional[Config] = None):
        """
        Initialize the MarkItDown converter.

        Args:
            config: Configuration for conversion options
        """
        super().__init__(config)
        self._markitdown = None
        self._markitdown_available = False
        self._openai_client = None

        # Try to import and initialize MarkItDown
        try:
            from markitdown import MarkItDown

            # Initialize MarkItDown with optional features
            llm_client = None
            llm_model = None
            llm_prompt = None
            docintel_endpoint = None
            docintel_key = None
            enable_plugins = getattr(config, 'markitdown_enable_plugins', False)

            # LLM configuration
            if getattr(config, 'llm_enabled', False):
                try:
                    from openai import OpenAI
                    api_key = os.getenv('OPENAI_API_KEY')
                    if api_key and api_key != 'your_openai_api_key_here':
                        self._openai_client = OpenAI(api_key=api_key)
                        llm_client = self._openai_client
                        llm_model = getattr(config, 'llm_model', 'gpt-4o')
                        llm_prompt = getattr(config, 'llm_prompt', None)
                except ImportError:
                    pass

            # Azure Document Intelligence configuration
            if getattr(config, 'azure_enabled', False):
                docintel_endpoint = os.getenv('AZURE_DOCINTEL_ENDPOINT')
                docintel_key = os.getenv('AZURE_DOCINTEL_KEY')
                if docintel_endpoint and docintel_endpoint != 'your_azure_endpoint_here':
                    # Azure endpoint is valid
                    pass
                else:
                    docintel_endpoint = None

            # Initialize MarkItDown with configuration
            self._markitdown = MarkItDown(
                llm_client=llm_client,
                llm_model=llm_model,
                llm_prompt=llm_prompt,
                docintel_endpoint=docintel_endpoint,
                enable_plugins=enable_plugins,
            )

            self._markitdown_available = True

        except ImportError:
            self._markitdown_available = False

    def convert(self, file_path: Path) -> ConversionResult:
        """
        Convert a document to Markdown using MarkItDown.

        Args:
            file_path: Path to the document file

        Returns:
            ConversionResult with markdown and optional rich metadata

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is not supported
            Exception: For conversion errors
        """
        # Validate file
        self.validate_file(file_path)

        if not self._markitdown:
            raise RuntimeError("MarkItDown is not available. Install with: pip install markitdown[all]")

        # Track conversion time
        start_time = time.time()

        try:
            # Convert using MarkItDown
            # Check if it's a YouTube URL (special case)
            file_str = str(file_path)
            if file_str.startswith('http://') or file_str.startswith('https://'):
                # URL conversion
                result = self._markitdown.convert(file_str)
            else:
                # File conversion
                result = self._markitdown.convert(str(file_path))

            # Extract markdown text
            markdown_text = result.text_content if hasattr(result, 'text_content') else str(result)

            # Calculate conversion time
            conversion_time = time.time() - start_time

            # Create metadata
            metadata = self._create_metadata(
                file_path=file_path,
                conversion_time=conversion_time,
                markdown_length=len(markdown_text)
            )

            # Create ConversionResult
            conversion_result = ConversionResult(
                markdown=markdown_text,
                images=[],  # MarkItDown doesn't extract images separately
                tables=[],  # MarkItDown doesn't extract tables separately
                metadata=metadata,
            )

            # Optional: Extract rich metadata if enabled
            if getattr(self.config, 'rich_conversion', False):
                conversion_result = self._enrich_result(conversion_result, file_path)

            return conversion_result

        except Exception as e:
            raise Exception(f"MarkItDown conversion failed: {str(e)}")

    def _create_metadata(
        self,
        file_path: Path,
        conversion_time: float,
        markdown_length: int
    ) -> ConversionMetadata:
        """
        Create basic conversion metadata.

        Args:
            file_path: Source file path
            conversion_time: Time taken for conversion
            markdown_length: Length of generated markdown

        Returns:
            ConversionMetadata object
        """
        file_size = file_path.stat().st_size if file_path.exists() else 0

        # Estimate word count (rough approximation)
        word_count = markdown_length // 5  # Average word length ~5 characters

        return ConversionMetadata(
            source_path=file_path,
            source_size_bytes=file_size,
            page_count=1,  # MarkItDown doesn't provide page count
            total_words=word_count,
            total_images=0,
            total_tables=0,
            strategy_used=self.config.strategy.value if hasattr(self.config.strategy, 'value') else str(self.config.strategy),
            converter_name=self.get_name(),
            conversion_time_seconds=conversion_time,
            timestamp=datetime.now(),
            warnings=[],
            errors=[],
        )

    def _enrich_result(
        self,
        result: ConversionResult,
        file_path: Path
    ) -> ConversionResult:
        """
        Enrich the conversion result with detailed metadata (optional).

        Uses existing libraries (PyMuPDF, python-docx, openpyxl) to extract
        detailed metadata, images, and tables.

        Args:
            result: Basic conversion result
            file_path: Source file path

        Returns:
            Enriched ConversionResult
        """
        extension = file_path.suffix.lower()

        try:
            # PDF rich metadata extraction
            if extension == '.pdf':
                result = self._enrich_pdf(result, file_path)

            # DOCX rich metadata extraction
            elif extension in ['.docx', '.doc']:
                result = self._enrich_docx(result, file_path)

            # XLSX rich metadata extraction
            elif extension in ['.xlsx', '.xls']:
                result = self._enrich_xlsx(result, file_path)

        except Exception as e:
            # If enrichment fails, just add a warning
            if result.metadata:
                result.metadata.warnings.append(f"Rich metadata extraction failed: {str(e)}")

        return result

    def _enrich_pdf(self, result: ConversionResult, file_path: Path) -> ConversionResult:
        """Extract rich metadata from PDF files using PyMuPDF."""
        try:
            import pymupdf as fitz

            doc = fitz.open(file_path)

            if result.metadata:
                result.metadata.page_count = len(doc)
                result.metadata.pdf_version = doc.metadata.get('format', 'PDF')
                result.metadata.title = doc.metadata.get('title')
                result.metadata.author = doc.metadata.get('author')
                result.metadata.subject = doc.metadata.get('subject')
                result.metadata.creator = doc.metadata.get('creator')
                result.metadata.producer = doc.metadata.get('producer')

            doc.close()

        except ImportError:
            pass  # PyMuPDF not available

        return result

    def _enrich_docx(self, result: ConversionResult, file_path: Path) -> ConversionResult:
        """Extract rich metadata from DOCX files using python-docx."""
        try:
            from docx import Document

            doc = Document(file_path)

            if result.metadata:
                result.metadata.title = doc.core_properties.title
                result.metadata.author = doc.core_properties.author
                result.metadata.subject = doc.core_properties.subject
                result.metadata.creator = doc.core_properties.last_modified_by
                result.metadata.creation_date = doc.core_properties.created
                result.metadata.modification_date = doc.core_properties.modified

                # Count paragraphs, tables, images
                result.metadata.total_tables = len(doc.tables)

                # Count inline shapes (images)
                image_count = 0
                for shape in doc.inline_shapes:
                    image_count += 1
                result.metadata.total_images = image_count

        except ImportError:
            pass  # python-docx not available

        return result

    def _enrich_xlsx(self, result: ConversionResult, file_path: Path) -> ConversionResult:
        """Extract rich metadata from XLSX files using openpyxl."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)

            if result.metadata:
                result.metadata.page_count = len(wb.sheetnames)  # Sheets as "pages"

                # Try to get document properties
                props = wb.properties
                if props:
                    result.metadata.title = props.title
                    result.metadata.author = props.creator
                    result.metadata.subject = props.subject
                    result.metadata.creation_date = props.created
                    result.metadata.modification_date = props.modified

            wb.close()

        except ImportError:
            pass  # openpyxl not available

        return result

    def supports_ocr(self) -> bool:
        """
        Check if OCR is supported.

        MarkItDown supports OCR for images through its built-in functionality.

        Returns:
            True (MarkItDown has OCR capabilities)
        """
        return True

    def get_name(self) -> str:
        """
        Get the converter name.

        Returns:
            Converter name
        """
        features = []
        if getattr(self.config, 'llm_enabled', False):
            features.append("LLM")
        if getattr(self.config, 'azure_enabled', False):
            features.append("Azure")

        if features:
            return f"MarkItDown Converter ({', '.join(features)})"
        return "MarkItDown Converter"

    def is_available(self) -> bool:
        """
        Check if MarkItDown is available.

        Returns:
            True if MarkItDown can be used
        """
        return self._markitdown_available

    def get_supported_extensions(self) -> List[str]:
        """
        Get list of supported file extensions.

        Returns:
            List of supported extensions
        """
        return self.SUPPORTED_EXTENSIONS

    def supports_youtube_urls(self) -> bool:
        """
        Check if YouTube URL conversion is supported.

        Returns:
            True if YouTube transcription is available
        """
        try:
            import youtube_transcript_api
            return True
        except ImportError:
            return False

    def estimate_conversion_time(self, file_path: Path) -> float:
        """
        Estimate conversion time based on file type and size.

        Args:
            file_path: Path to the file

        Returns:
            Estimated time in seconds
        """
        extension = file_path.suffix.lower()
        file_size_mb = file_path.stat().st_size / (1024 * 1024)

        # Different formats have different processing speeds
        if extension in ['.pdf']:
            # PDFs are relatively fast with MarkItDown
            return max(1.0, file_size_mb * 0.5)
        elif extension in ['.docx', '.xlsx', '.pptx']:
            # Office formats are fast
            return max(1.0, file_size_mb * 0.3)
        elif extension in ['.wav', '.mp3', '.m4a']:
            # Audio transcription is slower
            return max(5.0, file_size_mb * 2.0)
        elif extension in ['.jpg', '.jpeg', '.png']:
            # Images with OCR are moderate
            return max(2.0, file_size_mb * 1.0)
        else:
            # Default estimate
            return max(1.0, file_size_mb * 0.5)
